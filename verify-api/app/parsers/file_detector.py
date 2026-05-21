import io
import csv
from typing import BinaryIO

MYOB_GL_HEADERS = {"account", "job", "memo", "debit", "credit"}
MYOB_PL_HEADERS = {"account", "ytd", "this month"}
BANK_CSV_HEADERS = {"date", "description", "debit", "credit", "balance"}
QB_PL_KEYWORDS = {"profit & loss", "profit and loss", "income statement"}
BALANCE_SHEET_KEYWORDS = {"balance sheet", "statement of financial position"}
AGED_DEBTORS_KEYWORDS = {"aged debtors", "accounts receivable aging", "receivables aging"}
INVENTORY_KEYWORDS = {"inventory", "stock on hand", "item list"}
CUSTOMER_SALES_KEYWORDS = {"customer sales", "sales by customer", "customer report"}


def detect_file_type(file: BinaryIO, filename: str) -> str:
    header = file.read(8)
    file.seek(0)

    fname = filename.lower()

    # PDF
    if header[:4] == b"%PDF":
        text_sample = _pdf_text_sample(file)
        file.seek(0)
        tl = text_sample.lower()
        if any(k in tl for k in BALANCE_SHEET_KEYWORDS | QB_PL_KEYWORDS | AGED_DEBTORS_KEYWORDS):
            return "pdf_financial_report"
        return "bank_pdf"

    # Excel legacy
    if header[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" or fname.endswith(".xls"):
        return "xls_legacy"

    # Excel modern (ZIP-based .xlsx)
    if header[:4] == b"PK\x03\x04" or fname.endswith(".xlsx"):
        return "xlsx"

    if fname.endswith(".csv"):
        text = file.read(1024).decode("utf-8", errors="replace")
        file.seek(0)
        return _detect_csv_type(text)

    return "unknown"


def _detect_csv_type(text: str) -> str:
    tl = text.lower()
    first_lines = tl[:512]

    if any(k in first_lines for k in BALANCE_SHEET_KEYWORDS):
        return "balance_sheet"
    if any(k in first_lines for k in QB_PL_KEYWORDS):
        return "quickbooks_pl"
    if any(k in first_lines for k in AGED_DEBTORS_KEYWORDS):
        return "aged_debtors"
    if any(k in first_lines for k in INVENTORY_KEYWORDS):
        return "inventory"
    if any(k in first_lines for k in CUSTOMER_SALES_KEYWORDS):
        return "customer_sales"

    first_line = tl.split("\n")[0]
    cols = {c.strip().strip('"').strip("'") for c in first_line.split(",")}

    if MYOB_GL_HEADERS.issubset(cols):
        return "myob_gl"
    if MYOB_PL_HEADERS.issubset(cols):
        return "myob_pl"
    if BANK_CSV_HEADERS.issubset(cols):
        return "bank_csv"
    if "date" in cols and any(c in cols for c in ("description", "narrative", "memo", "particulars", "details")):
        return "bank_csv"

    return "csv_unknown"


def xlsx_to_csv_sheets(raw: bytes) -> list[tuple[str, str]]:
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([("" if v is None else str(v)) for v in row])
        csv_text = buf.getvalue()
        if csv_text.strip():
            results.append((sheet_name, csv_text))
    return results


def _pdf_text_sample(file: BinaryIO) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(file) as pdf:
            page = pdf.pages[0] if pdf.pages else None
            return page.extract_text() or "" if page else ""
    except Exception:
        return ""
